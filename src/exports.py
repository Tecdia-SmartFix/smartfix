"""XLSX + PDF export builders for Shift Logs and Fleet Analytics.

Each builder returns raw bytes so the FastAPI endpoint can hand them straight to
the client. Stylings are intentionally restrained: bold header rows, alternating
backgrounds, colored severity pills — readable for a non-technical manager.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ── Shared palette ────────────────────────────────────────────────────────────

BRAND      = "#1A53A1"
HEADER_BG  = "5A72A0"   # navy header row
HEADER_FG  = "FFFFFF"
ZEBRA_BG   = "F4F8FC"
BORDER_HEX = "B0B0B0"

SEVERITY_FILLS = {
    1: ("E5EEE4", "2E4E40", "Info"),
    2: ("FBFAF5", "6D5335", "Minor"),
    3: ("F6F4E8", "844D4D", "Degraded"),
    4: ("DC9B9B", "FFFFFF", "Impact"),
    5: ("F4D2D2", "4A1515", "Safety"),
}

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def _shift_for(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except (AttributeError, ValueError):
        return "—"
    h = d.hour
    return "1st shift" if 6 <= h < 18 else "2nd shift"


def _fmt_date(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except (AttributeError, ValueError):
        return "—"
    return f"{d.day:02d}-{MONTHS[d.month - 1]}-{d.year}"


def _fmt_time(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except (AttributeError, ValueError):
        return "—"
    return d.strftime("%H:%M")


def _phase_label(phase: Optional[str]) -> str:
    return {"start": "Pre-shift", "end": "End of shift"}.get(phase or "", phase or "—")


def _filter_summary(filters: dict) -> str:
    """Human-readable one-liner describing the active filters."""
    parts: list[str] = []
    if filters.get("machine_id"):
        parts.append(f"Machine: {filters['machine_id'].replace('_', ' ').title()}")
    if filters.get("category"):
        parts.append(f"Category: {filters['category']}")
    if filters.get("severity"):
        parts.append(f"Severity: {filters['severity']}")
    if filters.get("shift"):
        parts.append(f"Shift: {filters['shift']}")
    if filters.get("action_prefix"):
        parts.append(f"Action: {filters['action_prefix']}*")
    if filters.get("actor"):
        parts.append(f"Actor: {filters['actor']}")
    if filters.get("status"):
        parts.append(f"Status: {filters['status']}")
    if filters.get("days"):
        d = filters["days"]
        parts.append(f"Last {d} day{'s' if int(d) != 1 else ''}")
    if filters.get("date_from"):
        parts.append(f"From {filters['date_from']}")
    if filters.get("date_to"):
        parts.append(f"To {filters['date_to']}")
    return " · ".join(parts) if parts else "All data (no filters)"


def _machine_name(mid: str, name_map: dict) -> str:
    return name_map.get(mid) or (mid or "").replace("_", " ").title() or "—"


# ─────────────────────────────────────────────────────────────────────────────
# Shift logs — XLSX
# ─────────────────────────────────────────────────────────────────────────────

def shift_logs_xlsx(logs: list[dict], filters: dict, machine_names: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Logs"

    thin = Side(border_style="thin", color=BORDER_HEX)
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws["A1"] = "Shift Logs Export"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="0F1C3F")
    ws.merge_cells("A1:I1")

    # Subtitle: filters + generated timestamp
    ws["A2"] = f"{_filter_summary(filters)}  ·  Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="5A72A0", italic=True)
    ws.merge_cells("A2:I2")

    # Stats strip
    total = len(logs)
    issues = sum(len(l.get("anomalies") or []) for l in logs)
    crit   = sum(1 for l in logs if (l.get("severity") or 0) >= 4)
    ws["A3"] = f"Total logs: {total}    Issues found: {issues}    Critical alerts: {crit}"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="1A53A1")
    ws.merge_cells("A3:I3")
    ws.row_dimensions[3].height = 22

    ws.append([])  # spacer row 4

    headers = ["Date", "Time", "Shift", "Phase", "Machine", "Worker", "Severity", "Anomalies", "Notes"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = box
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Data rows
    for i, log in enumerate(logs):
        sev = log.get("severity") or 1
        sev_bg, sev_fg, sev_word = SEVERITY_FILLS.get(sev, SEVERITY_FILLS[1])
        anomalies = ", ".join(a.get("title", "") for a in (log.get("anomalies") or [])) or "—"
        row_vals = [
            _fmt_date(log.get("created_at")),
            _fmt_time(log.get("created_at")),
            _shift_for(log.get("created_at")),
            "Void" if log.get("void_at") else _phase_label(log.get("phase")),
            _machine_name(log.get("machine_id") or "", machine_names),
            log.get("worker_label") or "—",
            f"{sev} — {sev_word}",
            anomalies,
            log.get("notes") or "—",
        ]
        ws.append(row_vals)
        r = ws.max_row
        zebra = (i % 2 == 1)
        for col_idx in range(1, len(row_vals) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.font = Font(name="Calibri", size=10, color="2E4E40")
            c.alignment = Alignment(vertical="center", wrap_text=(col_idx in (8, 9)))
            c.border = box
            if zebra:
                c.fill = PatternFill("solid", fgColor=ZEBRA_BG)
        # Severity pill
        sc = ws.cell(row=r, column=7)
        sc.font = Font(name="Calibri", size=10, bold=True, color=sev_fg)
        sc.fill = PatternFill("solid", fgColor=sev_bg)
        sc.alignment = Alignment(vertical="center", horizontal="center")

    # Column widths
    widths = [13, 8, 10, 14, 28, 22, 14, 40, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Shift logs — PDF
# ─────────────────────────────────────────────────────────────────────────────

def _styles():
    base = getSampleStyleSheet()
    return {
        "title":   ParagraphStyle("title",   parent=base["Title"],   fontSize=22, leading=26, textColor=colors.HexColor("#0F1C3F"), spaceAfter=4),
        "sub":     ParagraphStyle("sub",     parent=base["Normal"],  fontSize=9,  leading=12, textColor=colors.HexColor("#5A72A0"), spaceAfter=2),
        "kpi":     ParagraphStyle("kpi",     parent=base["Normal"],  fontSize=11, leading=14, textColor=colors.HexColor("#1A53A1"), spaceAfter=10, fontName="Helvetica-Bold"),
        "section": ParagraphStyle("section", parent=base["Heading2"],fontSize=13, leading=16, textColor=colors.HexColor("#0F1C3F"), spaceBefore=14, spaceAfter=6),
        "cell":    ParagraphStyle("cell",    parent=base["Normal"],  fontSize=9,  leading=11, textColor=colors.HexColor("#2E4E40")),
        "cell_b":  ParagraphStyle("cell_b",  parent=base["Normal"],  fontSize=9,  leading=11, textColor=colors.HexColor("#0F1C3F"), fontName="Helvetica-Bold"),
    }


def _page_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6D7C74"))
    canvas.drawString(15 * mm, 10 * mm, "Tecdia SmartFix")
    canvas.drawRightString(doc.pagesize[0] - 15 * mm, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _table_with_header(data: list[list], col_widths: list[float], severity_col: Optional[int] = None) -> Table:
    """Build a styled table. `severity_col` is a 0-based index of a column whose
    cell value starts with the severity number, so we can tint the cell."""
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#5A72A0")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, 0), "LEFT"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING",    (0, 0), (-1, 0), 6),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7E2")),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F4F8FC")))
        if severity_col is not None:
            cell = data[i][severity_col]
            text = cell.text if hasattr(cell, "text") else str(cell)
            try:
                sev = int(text.strip().split()[0])
            except (ValueError, IndexError):
                sev = None
            if sev in SEVERITY_FILLS:
                bg, fg, _ = SEVERITY_FILLS[sev]
                style.append(("BACKGROUND", (severity_col, i), (severity_col, i), colors.HexColor(f"#{bg}")))
                style.append(("TEXTCOLOR",  (severity_col, i), (severity_col, i), colors.HexColor(f"#{fg}")))
                style.append(("FONTNAME",   (severity_col, i), (severity_col, i), "Helvetica-Bold"))
                style.append(("ALIGN",      (severity_col, i), (severity_col, i), "CENTER"))
    t.setStyle(TableStyle(style))
    return t


def shift_logs_pdf(logs: list[dict], filters: dict, machine_names: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=18 * mm,
        title="Shift Logs Export",
    )
    s = _styles()

    story: list = [
        Paragraph("Shift Logs Export", s["title"]),
        Paragraph(_filter_summary(filters), s["sub"]),
        Paragraph(f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}", s["sub"]),
        Spacer(1, 8),
    ]

    total = len(logs)
    issues = sum(len(l.get("anomalies") or []) for l in logs)
    crit   = sum(1 for l in logs if (l.get("severity") or 0) >= 4)
    story.append(Paragraph(
        f"Total logs: {total}     Issues found: {issues}     Critical alerts: {crit}",
        s["kpi"],
    ))

    headers = ["Date", "Time", "Shift", "Phase", "Machine", "Worker", "Severity", "Anomalies"]
    rows: list[list] = [headers]
    for log in logs:
        sev = log.get("severity") or 1
        _, _, sev_word = SEVERITY_FILLS.get(sev, SEVERITY_FILLS[1])
        anomalies = ", ".join(a.get("title", "") for a in (log.get("anomalies") or [])) or "—"
        rows.append([
            _fmt_date(log.get("created_at")),
            _fmt_time(log.get("created_at")),
            _shift_for(log.get("created_at")),
            "Void" if log.get("void_at") else _phase_label(log.get("phase")),
            Paragraph(_machine_name(log.get("machine_id") or "", machine_names), s["cell_b"]),
            log.get("worker_label") or "—",
            f"{sev} {sev_word}",
            Paragraph(anomalies, s["cell"]),
        ])

    # Column widths chosen to fit A4 landscape minus margins (~267mm usable).
    col_widths = [22 * mm, 16 * mm, 22 * mm, 26 * mm, 50 * mm, 36 * mm, 26 * mm, 65 * mm]
    story.append(_table_with_header(rows, col_widths, severity_col=6))

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Analytics — XLSX (multi-sheet)
# ─────────────────────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, ncols: int):
    thin = Side(border_style="thin", color=BORDER_HEX)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = box
    ws.row_dimensions[row].height = 24


def _zebra_rows(ws, start_row: int, end_row: int, ncols: int):
    thin = Side(border_style="thin", color=BORDER_HEX)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Calibri", size=10, color="2E4E40")
            c.alignment = Alignment(vertical="center")
            c.border = box
            if (r - start_row) % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ZEBRA_BG)


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_sheet(wb, name: str, headers: list[str], rows: Iterable[list], widths: list[int],
                 title: str, subtitle: str):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="0F1C3F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color="5A72A0")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, ws.max_row, len(headers))
    header_row = ws.max_row
    start = header_row + 1
    for row in rows:
        ws.append(row)
    _zebra_rows(ws, start, ws.max_row, len(headers))
    _set_widths(ws, widths)
    ws.freeze_panes = ws.cell(row=start, column=1)
    return ws


def analytics_xlsx(data: dict, filters: dict) -> bytes:
    wb = Workbook()
    # Replace the default sheet with our "Summary" sheet.
    default = wb.active
    wb.remove(default)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Fleet Analytics Export"
    summary["A1"].font = Font(size=18, bold=True, color="0F1C3F")
    summary.merge_cells("A1:D1")
    summary["A2"] = f"{_filter_summary(filters)}  ·  Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    summary["A2"].font = Font(size=10, italic=True, color="5A72A0")
    summary.merge_cells("A2:D2")
    summary.append([])

    totals = data.get("totals", {})
    summary.append(["Metric", "Value"])
    _style_header_row(summary, summary.max_row, 2)
    start = summary.max_row + 1
    summary.append(["Total queries",    totals.get("queries", 0)])
    summary.append(["Alerts fired",     totals.get("alerts", 0)])
    summary.append(["Active machines",  totals.get("machines", 0)])
    summary.append(["Alert rate (%)",   totals.get("alert_rate_pct", 0)])
    _zebra_rows(summary, start, summary.max_row, 2)
    _set_widths(summary, [28, 18])

    sub = _filter_summary(filters)

    _build_sheet(
        wb, "Per Machine",
        ["Machine", "Queries", "Alerts", "Alert rate (%)", "Avg severity"],
        [
            [m["display_name"], m["query_count"], m["alert_count"], m["alert_rate_pct"], m["avg_severity"]]
            for m in data.get("per_machine", [])
        ],
        widths=[32, 12, 12, 16, 14],
        title="Per-Machine Activity", subtitle=sub,
    )

    _build_sheet(
        wb, "Code Frequency",
        ["Code", "Machine", "Count", "Avg severity"],
        [[c["code"], c["machine"], c["count"], c.get("avg_severity", 0)] for c in data.get("code_frequency", [])],
        widths=[18, 32, 12, 14],
        title="Top Error / Question Codes", subtitle=sub,
    )

    sev_dist = data.get("severity_distribution", {})
    _, _, _ = SEVERITY_FILLS[1]
    _build_sheet(
        wb, "Severity",
        ["Severity", "Label", "Count"],
        [[k, SEVERITY_FILLS[int(k)][2], v] for k, v in sev_dist.items()],
        widths=[12, 16, 12],
        title="Severity Distribution", subtitle=sub,
    )

    _build_sheet(
        wb, "Hourly (24h)",
        ["Hour", "Queries"],
        [[b["hour"], b["count"]] for b in data.get("queries_per_hour_24h", [])],
        widths=[10, 12],
        title="Queries per Hour (last 24h)", subtitle="Hour bands are facility-local · ignores filters",
    )

    _build_sheet(
        wb, "Top Questions",
        ["Question", "Machine", "Count"],
        [[q["question"], q["machine"], q["count"]] for q in data.get("top_questions", [])],
        widths=[60, 30, 12],
        title="Top Verbatim Questions", subtitle=sub,
    )

    _build_sheet(
        wb, "Failure Likelihood",
        ["Machine", "Score", "Risk"],
        [[f["display_name"], f.get("score", 0), f.get("risk", "—")] for f in data.get("failure_likelihood", [])],
        widths=[32, 12, 18],
        title="Failure Likelihood", subtitle="Per-asset; ignores filters",
    )

    _build_sheet(
        wb, "Depreciation",
        ["Machine", "Age (yrs)", "Remaining (%)", "Status"],
        [[d["display_name"], d.get("age_years", "—"), d.get("remaining_pct", "—"), d.get("status", "—")]
         for d in data.get("depreciation", [])],
        widths=[32, 12, 16, 18],
        title="Depreciation", subtitle="Per-asset; ignores filters",
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Analytics — PDF (multi-section)
# ─────────────────────────────────────────────────────────────────────────────

def analytics_pdf(data: dict, filters: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=18 * mm,
        title="Fleet Analytics Export",
    )
    s = _styles()
    story: list = [
        Paragraph("Fleet Analytics Export", s["title"]),
        Paragraph(_filter_summary(filters), s["sub"]),
        Paragraph(f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}", s["sub"]),
        Spacer(1, 8),
    ]

    totals = data.get("totals", {})
    story.append(Paragraph(
        f"Queries: {totals.get('queries', 0)}     Alerts: {totals.get('alerts', 0)}     "
        f"Machines: {totals.get('machines', 0)}     Alert rate: {totals.get('alert_rate_pct', 0)}%",
        s["kpi"],
    ))

    usable_w = doc.width  # portrait A4 minus margins ≈ 174mm

    def add_section(heading: str, headers: list[str], rows: list[list], col_widths: list[float]):
        if not rows:
            return
        story.append(Paragraph(heading, s["section"]))
        story.append(_table_with_header([headers] + rows, col_widths))

    add_section(
        "Per-Machine Activity",
        ["Machine", "Queries", "Alerts", "Alert rate", "Avg severity"],
        [[Paragraph(m["display_name"], s["cell_b"]), m["query_count"], m["alert_count"],
          f"{m['alert_rate_pct']}%", m["avg_severity"]]
         for m in data.get("per_machine", [])],
        [usable_w * 0.42, usable_w * 0.13, usable_w * 0.13, usable_w * 0.16, usable_w * 0.16],
    )

    add_section(
        "Top Error / Question Codes",
        ["Code", "Machine", "Count", "Avg severity"],
        [[c["code"], Paragraph(c["machine"], s["cell"]), c["count"], c.get("avg_severity", 0)]
         for c in data.get("code_frequency", [])],
        [usable_w * 0.18, usable_w * 0.50, usable_w * 0.14, usable_w * 0.18],
    )

    sev_dist = data.get("severity_distribution", {})
    add_section(
        "Severity Distribution",
        ["Severity", "Label", "Count"],
        [[k, SEVERITY_FILLS[int(k)][2], v] for k, v in sev_dist.items()],
        [usable_w * 0.20, usable_w * 0.50, usable_w * 0.30],
    )

    add_section(
        "Top Verbatim Questions",
        ["Question", "Machine", "Count"],
        [[Paragraph(q["question"], s["cell"]), Paragraph(q["machine"], s["cell"]), q["count"]]
         for q in data.get("top_questions", [])],
        [usable_w * 0.55, usable_w * 0.30, usable_w * 0.15],
    )

    add_section(
        "Failure Likelihood",
        ["Machine", "Score", "Risk"],
        [[Paragraph(f["display_name"], s["cell_b"]), f.get("score", 0), f.get("risk", "—")]
         for f in data.get("failure_likelihood", [])],
        [usable_w * 0.55, usable_w * 0.20, usable_w * 0.25],
    )

    add_section(
        "Depreciation",
        ["Machine", "Age (yrs)", "Remaining", "Status"],
        [[Paragraph(d["display_name"], s["cell_b"]), d.get("age_years", "—"),
          f"{d.get('remaining_pct', '—')}%", d.get("status", "—")]
         for d in data.get("depreciation", [])],
        [usable_w * 0.42, usable_w * 0.15, usable_w * 0.20, usable_w * 0.23],
    )

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Audit log — XLSX
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_audit_ts(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except (AttributeError, ValueError):
        return iso
    return d.strftime("%d %b %Y · %H:%M:%S")


def audit_logs_xlsx(entries: list[dict], filters: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    thin = Side(border_style="thin", color=BORDER_HEX)
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Audit Log Export"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="0F1C3F")
    ws.merge_cells("A1:G1")

    ws["A2"] = f"{_filter_summary(filters)}  ·  Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="5A72A0", italic=True)
    ws.merge_cells("A2:G2")

    total = len(entries)
    success = sum(1 for e in entries if (e.get("status") or "").lower() == "success")
    failures = total - success
    ws["A3"] = f"Total events: {total}    Successes: {success}    Failures: {failures}"
    ws["A3"].font = Font(name="Calibri", size=11, bold=True, color="1A53A1")
    ws.merge_cells("A3:G3")
    ws.row_dimensions[3].height = 22

    ws.append([])

    headers = ["Timestamp", "Action", "Actor", "Target", "Status", "IP", "Details"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = box
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    fail_fill = PatternFill("solid", fgColor="FDE7E7")
    fail_fg   = "B91C1C"
    ok_fill   = PatternFill("solid", fgColor="ECFDF5")
    ok_fg     = "047857"

    for i, e in enumerate(entries):
        is_success = (e.get("status") or "").lower() == "success"
        details = e.get("details") or {}
        details_str = ", ".join(f"{k}={v}" for k, v in details.items()) if isinstance(details, dict) else str(details)
        row_vals = [
            _fmt_audit_ts(e.get("ts")),
            e.get("action") or "—",
            e.get("actor") or "—",
            e.get("target") or "—",
            (e.get("status") or "—").title(),
            e.get("ip") or "—",
            details_str or "—",
        ]
        ws.append(row_vals)
        r = ws.max_row
        zebra = (i % 2 == 1)
        for col_idx in range(1, len(row_vals) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.font = Font(name="Calibri", size=10, color="2E4E40")
            c.alignment = Alignment(vertical="center", wrap_text=(col_idx == 7))
            c.border = box
            if zebra:
                c.fill = PatternFill("solid", fgColor=ZEBRA_BG)
        # Status pill
        sc = ws.cell(row=r, column=5)
        sc.font = Font(name="Calibri", size=10, bold=True, color=(ok_fg if is_success else fail_fg))
        sc.fill = ok_fill if is_success else fail_fill
        sc.alignment = Alignment(vertical="center", horizontal="center")

    widths = [22, 30, 28, 28, 12, 16, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Audit log — PDF
# ─────────────────────────────────────────────────────────────────────────────

def audit_logs_pdf(entries: list[dict], filters: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=18 * mm,
        title="Audit Log Export",
    )
    s = _styles()

    story: list = [
        Paragraph("Audit Log Export", s["title"]),
        Paragraph(_filter_summary(filters), s["sub"]),
        Paragraph(f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}", s["sub"]),
        Spacer(1, 8),
    ]

    total = len(entries)
    success = sum(1 for e in entries if (e.get("status") or "").lower() == "success")
    failures = total - success
    story.append(Paragraph(
        f"Total events: {total}     Successes: {success}     Failures: {failures}",
        s["kpi"],
    ))

    headers = ["Timestamp", "Action", "Actor", "Target", "Status", "IP", "Details"]
    rows: list[list] = [headers]
    for e in entries:
        details = e.get("details") or {}
        details_str = ", ".join(f"{k}={v}" for k, v in details.items()) if isinstance(details, dict) else str(details)
        rows.append([
            _fmt_audit_ts(e.get("ts")),
            Paragraph(e.get("action") or "—", s["cell_b"]),
            Paragraph(e.get("actor") or "—", s["cell"]),
            Paragraph(e.get("target") or "—", s["cell"]),
            (e.get("status") or "—").title(),
            e.get("ip") or "—",
            Paragraph(details_str or "—", s["cell"]),
        ])

    # 277mm usable in landscape A4 minus margins.
    col_widths = [30 * mm, 38 * mm, 36 * mm, 36 * mm, 20 * mm, 24 * mm, 53 * mm]

    # Build the table directly so we can tint the Status column per-row.
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#5A72A0")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, 0), "LEFT"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING",    (0, 0), (-1, 0), 6),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7E2")),
    ]
    for i in range(1, len(rows)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F4F8FC")))
        is_success = str(rows[i][4]).lower() == "success"
        style.append((
            "BACKGROUND", (4, i), (4, i),
            colors.HexColor("#ECFDF5") if is_success else colors.HexColor("#FDE7E7"),
        ))
        style.append((
            "TEXTCOLOR", (4, i), (4, i),
            colors.HexColor("#047857") if is_success else colors.HexColor("#B91C1C"),
        ))
        style.append(("FONTNAME", (4, i), (4, i), "Helvetica-Bold"))
        style.append(("ALIGN",    (4, i), (4, i), "CENTER"))
    t.setStyle(TableStyle(style))
    story.append(t)

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    return buf.getvalue()
